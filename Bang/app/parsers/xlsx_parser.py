from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.core.ids import stable_id
from app.core.normalizers import normalize_entity_key, normalize_text, parse_quantity
from app.core.segments import ArtifactSegment
from app.core.schemas import (
    ArtifactType,
    AtomType,
    AuthorityClass,
    EvidenceAtom,
    ReviewStatus,
    SourceRef,
    ParserCapability,
    ParserMatch,
)
from app.parsers.base import BaseParser
from app.parsers.segmenters import segment_xlsx
from app.domain.schemas import DomainPack

HEADER_ALIASES = {
    "site": {"site", "location", "facility", "store", "building"},
    "device": {"device", "asset", "equipment", "camera", "ap", "reader", "item"},
    "quantity": {"qty", "qty.", "quantity", "count", "#"},
    "floor": {"floor", "level"},
    "room": {"room", "area", "zone"},
    "scope": {"scope", "included", "work type"},
    "access": {"access", "access window", "hours", "site access"},
}

SKIP_ROW_MARKERS = ("total", "subtotal", "grand total")


class XlsxParser(BaseParser):
    parser_name = "xlsx"
    parser_version = "xlsx_parser_v1"
    capability = ParserCapability(
        parser_name=parser_name,
        parser_version=parser_version,
        supported_extensions=[".xlsx", ".csv"],
        supported_artifact_types=[ArtifactType.xlsx, ArtifactType.csv],
        emitted_atom_types=[AtomType.entity, AtomType.quantity, AtomType.scope_item, AtomType.constraint],
        supported_domain_packs=["*"],
        requires_binary=False,
        supports_source_replay=True,
    )

    def match(self, path: Path, sample_text: str | None, domain_pack: DomainPack | None) -> ParserMatch:
        del sample_text, domain_pack
        suffix = path.suffix.lower()
        confidence = 0.0
        reasons: list[str] = []
        if suffix in {".xlsx", ".csv"}:
            confidence = 0.72
            reasons.append(f"spreadsheet_extension:{suffix}")
        return ParserMatch(
            parser_name=self.parser_name,
            confidence=confidence,
            reasons=reasons,
            artifact_type=ArtifactType.xlsx if suffix == ".xlsx" else ArtifactType.csv,
        )

    def parse(self, artifact_path: Path) -> list[EvidenceAtom]:
        artifact_id = stable_id("art", str(artifact_path))
        return self.parse_artifact(
            project_id="unknown_project",
            artifact_id=artifact_id,
            path=artifact_path,
        )

    def segment_artifact(self, project_id: str, artifact_id: str, path: Path) -> list[ArtifactSegment]:
        return segment_xlsx(project_id=project_id, artifact_id=artifact_id, path=path, parser_version=self.parser_version)

    def parse_artifact(
        self,
        project_id: str,
        artifact_id: str,
        path: Path,
        domain_pack: DomainPack | None = None,
    ) -> list[EvidenceAtom]:
        del domain_pack
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._parse_csv(project_id=project_id, artifact_id=artifact_id, path=path)
        return self._parse_xlsx(project_id=project_id, artifact_id=artifact_id, path=path)

    def _parse_xlsx(self, project_id: str, artifact_id: str, path: Path) -> list[EvidenceAtom]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        atoms: list[EvidenceAtom] = []
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            atoms.extend(
                self._parse_sheet_rows(
                    project_id=project_id,
                    artifact_id=artifact_id,
                    filename=path.name,
                    artifact_type=ArtifactType.xlsx,
                    sheet_name=sheet.title,
                    rows=rows,
                )
            )
        return atoms

    def _parse_csv(self, project_id: str, artifact_id: str, path: Path) -> list[EvidenceAtom]:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            rows = [list(row) for row in reader]
        return self._parse_sheet_rows(
            project_id=project_id,
            artifact_id=artifact_id,
            filename=path.name,
            artifact_type=ArtifactType.csv,
            sheet_name="csv",
            rows=rows,
        )

    def _parse_sheet_rows(
        self,
        project_id: str,
        artifact_id: str,
        filename: str,
        artifact_type: ArtifactType,
        sheet_name: str,
        rows: list[list[Any]],
    ) -> list[EvidenceAtom]:
        if not rows:
            return []
        header_idx, header_map = self._detect_header(rows)
        if header_idx is None:
            return []

        atoms: list[EvidenceAtom] = []
        for row_idx in range(header_idx + 1, len(rows)):
            row = rows[row_idx]
            if self._is_blank_row(row):
                continue
            if self._is_total_row(row):
                continue
            extracted = self._extract_row_values(row, header_map)
            if self._is_blank_row(list(extracted.values())):
                continue
            atoms.extend(
                self._row_to_atoms(
                    project_id=project_id,
                    artifact_id=artifact_id,
                    artifact_type=artifact_type,
                    filename=filename,
                    sheet_name=sheet_name,
                    row_number=row_idx + 1,
                    header_map=header_map,
                    extracted=extracted,
                )
            )
        return atoms

    def _detect_header(self, rows: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
        scan_limit = min(25, len(rows))
        best_idx: int | None = None
        best_map: dict[str, int] = {}
        best_score = -1

        for idx in range(scan_limit):
            row = rows[idx]
            current_map: dict[str, int] = {}
            for col_idx, cell in enumerate(row):
                cell_text = normalize_text(str(cell or "")).strip(".:")
                if not cell_text:
                    continue
                for canonical, aliases in HEADER_ALIASES.items():
                    if cell_text in aliases and canonical not in current_map:
                        current_map[canonical] = col_idx
            score = len(current_map)
            if score > best_score:
                best_score = score
                best_idx = idx
                best_map = current_map

        if best_score <= 0:
            return None, {}
        return best_idx, best_map

    def _extract_row_values(self, row: list[Any], header_map: dict[str, int]) -> dict[str, str]:
        extracted: dict[str, str] = {}
        for key, idx in header_map.items():
            value = row[idx] if idx < len(row) else ""
            extracted[key] = str(value).strip() if value is not None else ""
        return extracted

    def _is_blank_row(self, row: list[Any]) -> bool:
        return all(str(cell or "").strip() == "" for cell in row)

    def _is_total_row(self, row: list[Any]) -> bool:
        joined = " ".join(str(cell or "").strip().lower() for cell in row)
        return any(marker in joined for marker in SKIP_ROW_MARKERS)

    def _build_entity_keys(self, extracted: dict[str, str]) -> list[str]:
        keys: list[str] = []
        for entity_type in ("site", "device", "floor", "room"):
            value = extracted.get(entity_type, "").strip()
            if value:
                keys.append(normalize_entity_key(entity_type, value))
        return keys

    def _build_source_ref(
        self,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        row_number: int,
        header_map: dict[str, int],
    ) -> SourceRef:
        columns = {key: get_column_letter(index + 1) for key, index in header_map.items()}
        return SourceRef(
            id=stable_id("src", artifact_id, sheet_name, row_number),
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            locator={"sheet": sheet_name, "row": row_number, "columns": columns},
            extraction_method="table_header_mapping",
            parser_version=self.parser_version,
        )

    def _row_confidence(self, extracted: dict[str, str]) -> float:
        major_count = sum(1 for key in ("site", "device", "quantity") if extracted.get(key))
        return 0.92 if major_count == 3 else 0.8

    def _row_to_atoms(
        self,
        project_id: str,
        artifact_id: str,
        artifact_type: ArtifactType,
        filename: str,
        sheet_name: str,
        row_number: int,
        header_map: dict[str, int],
        extracted: dict[str, str],
    ) -> list[EvidenceAtom]:
        entity_keys = self._build_entity_keys(extracted)
        row_confidence = self._row_confidence(extracted)
        source_ref = self._build_source_ref(
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            sheet_name=sheet_name,
            row_number=row_number,
            header_map=header_map,
        )
        atoms: list[EvidenceAtom] = []

        def append_atom(
            atom_type: AtomType,
            raw_text: str,
            value: dict[str, Any],
            review_status: ReviewStatus = ReviewStatus.auto_accepted,
            review_flags: list[str] | None = None,
        ) -> None:
            atoms.append(
                EvidenceAtom(
                    id=stable_id("atm", project_id, artifact_id, sheet_name, row_number, atom_type.value, raw_text),
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=atom_type,
                    raw_text=raw_text,
                    normalized_text=normalize_text(raw_text),
                    value=value,
                    entity_keys=entity_keys,
                    source_refs=[source_ref],
                    authority_class=AuthorityClass.approved_site_roster,
                    confidence=row_confidence,
                    review_status=review_status,
                    review_flags=review_flags or [],
                    parser_version=self.parser_version,
                )
            )

        site = extracted.get("site", "").strip()
        device = extracted.get("device", "").strip()
        quantity_raw = extracted.get("quantity", "").strip()
        scope = extracted.get("scope", "").strip()
        access = extracted.get("access", "").strip()
        floor = extracted.get("floor", "").strip()
        room = extracted.get("room", "").strip()

        if site:
            append_atom(AtomType.entity, f"Site {site}", {"entity_type": "site", "name": site})
        if device:
            append_atom(AtomType.entity, f"Device {device}", {"entity_type": "device", "name": device})
        if quantity_raw:
            parsed_quantity = parse_quantity(quantity_raw)
            review_status = (
                ReviewStatus.needs_review if parsed_quantity.get("uncertain") else ReviewStatus.auto_accepted
            )
            flags = ["quantity_uncertain"] if parsed_quantity.get("uncertain") else []
            append_atom(
                AtomType.quantity,
                f"Quantity {quantity_raw}",
                parsed_quantity,
                review_status=review_status,
                review_flags=flags,
            )

        if scope or (site and device):
            work_scope = scope if scope else "work_item"
            append_atom(
                AtomType.scope_item,
                f"Scope {work_scope}",
                {
                    "scope": work_scope,
                    "site": site,
                    "device": device,
                    "floor": floor,
                    "room": room,
                },
            )

        if access:
            append_atom(
                AtomType.constraint,
                f"Access {access}",
                {"access_window": access, "site": site, "device": device},
            )

        return atoms
