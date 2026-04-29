from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.core.ids import stable_id
from app.core.normalizers import normalize_entity_key, normalize_text, parse_quantity
from app.core.segments import ArtifactSegment
from app.core.schemas import (
    ArtifactType,
    AtomType,
    AuthorityClass,
    EvidenceAtom,
    ReviewStatus,
    SourceRef,
    ParserCapability,
    ParserMatch,
)
from app.parsers.base import BaseParser
from app.parsers.segmenters import segment_quote
from app.domain.schemas import DomainPack

HEADER_ALIASES = {
    "part_number": {"part", "part number", "sku", "item number"},
    "description": {"description", "item", "product", "device"},
    "quantity": {"qty", "quantity", "count"},
    "unit_price": {"unit price", "price", "cost"},
    "lead_time": {"lead time", "eta"},
}


class QuoteParser(BaseParser):
    parser_name = "quote"
    parser_version = "quote_parser_v1"
    capability = ParserCapability(
        parser_name=parser_name,
        parser_version=parser_version,
        supported_extensions=[".xlsx", ".csv", ".txt"],
        supported_artifact_types=[ArtifactType.vendor_quote, ArtifactType.xlsx, ArtifactType.csv, ArtifactType.txt],
        emitted_atom_types=[AtomType.vendor_line_item, AtomType.quantity, AtomType.constraint],
        supported_domain_packs=["*"],
        requires_binary=False,
        supports_source_replay=True,
    )

    def match(self, path: Path, sample_text: str | None, domain_pack: DomainPack | None) -> ParserMatch:
        del domain_pack
        suffix = path.suffix.lower()
        filename = path.name.lower()
        confidence = 0.0
        reasons: list[str] = []
        if suffix not in {".xlsx", ".csv", ".txt"}:
            return ParserMatch(
                parser_name=self.parser_name,
                confidence=0.0,
                reasons=[],
                artifact_type=ArtifactType.vendor_quote,
            )
        if any(token in filename for token in ("quote", "po", "vendor")):
            confidence = 0.95
            reasons.append("filename_quote_hint")
        elif self.looks_like_quote_artifact(path):
            confidence = 0.86
            reasons.append("header_quote_hint")
        elif sample_text and "part number" in normalize_text(sample_text):
            confidence = 0.8
            reasons.append("text_part_number_hint")
        return ParserMatch(
            parser_name=self.parser_name,
            confidence=confidence,
            reasons=reasons,
            artifact_type=ArtifactType.vendor_quote,
        )

    def parse(self, artifact_path: Path) -> list[EvidenceAtom]:
        artifact_id = stable_id("art", str(artifact_path))
        return self.parse_artifact("unknown_project", artifact_id, artifact_path)

    def segment_artifact(self, project_id: str, artifact_id: str, path: Path) -> list[ArtifactSegment]:
        return segment_quote(project_id=project_id, artifact_id=artifact_id, path=path, parser_version=self.parser_version)

    def parse_artifact(
        self,
        project_id: str,
        artifact_id: str,
        path: Path,
        domain_pack: DomainPack | None = None,
    ) -> list[EvidenceAtom]:
        del domain_pack
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            return self._parse_xlsx(project_id=project_id, artifact_id=artifact_id, path=path)
        if suffix == ".csv":
            return self._parse_csv(project_id=project_id, artifact_id=artifact_id, path=path)
        if suffix == ".txt":
            return self._parse_txt(project_id=project_id, artifact_id=artifact_id, path=path)
        return []

    @classmethod
    def looks_like_quote_artifact(cls, path: Path) -> bool:
        name = path.name.lower()
        if any(token in name for token in ("quote", "vendor", "po", "purchase_order")):
            return True

        suffix = path.suffix.lower()
        try:
            if suffix == ".xlsx":
                workbook = load_workbook(path, read_only=True, data_only=True)
                for sheet in workbook.worksheets:
                    sample_rows = [list(row) for _, row in zip(range(5), sheet.iter_rows(values_only=True))]
                    _, header_map = cls._detect_header(sample_rows)
                    if cls._is_quote_header_map(header_map):
                        return True
            elif suffix in {".csv", ".txt"}:
                content = path.read_text(encoding="utf-8", errors="ignore")
                sample_rows = [re.split(r"[,\t|]", line) for line in content.splitlines()[:5] if line.strip()]
                _, header_map = cls._detect_header(sample_rows)
                if cls._is_quote_header_map(header_map):
                    return True
        except Exception:
            return False
        return False

    @staticmethod
    def _is_quote_header_map(header_map: dict[str, int]) -> bool:
        has_core = {"description", "quantity"}.issubset(set(header_map.keys()))
        has_quote_specific = any(k in header_map for k in ("part_number", "unit_price", "lead_time"))
        return has_core and has_quote_specific

    def _parse_xlsx(self, project_id: str, artifact_id: str, path: Path) -> list[EvidenceAtom]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        atoms: list[EvidenceAtom] = []
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            atoms.extend(
                self._parse_rows(
                    project_id=project_id,
                    artifact_id=artifact_id,
                    filename=path.name,
                    sheet_name=sheet.title,
                    artifact_type=ArtifactType.xlsx,
                    rows=rows,
                )
            )
        return atoms

    def _parse_csv(self, project_id: str, artifact_id: str, path: Path) -> list[EvidenceAtom]:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            rows = [list(row) for row in reader]
        return self._parse_rows(
            project_id=project_id,
            artifact_id=artifact_id,
            filename=path.name,
            sheet_name="csv",
            artifact_type=ArtifactType.csv,
            rows=rows,
        )

    def _parse_txt(self, project_id: str, artifact_id: str, path: Path) -> list[EvidenceAtom]:
        content = path.read_text(encoding="utf-8", errors="ignore")
        lines = [line for line in content.splitlines() if line.strip()]
        rows = [re.split(r"[,\t|]", line) for line in lines]
        return self._parse_rows(
            project_id=project_id,
            artifact_id=artifact_id,
            filename=path.name,
            sheet_name="txt",
            artifact_type=ArtifactType.txt,
            rows=rows,
        )

    @classmethod
    def _detect_header(cls, rows: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
        scan_limit = min(25, len(rows))
        best_idx: int | None = None
        best_map: dict[str, int] = {}
        best_score = -1

        for idx in range(scan_limit):
            row = rows[idx]
            current_map: dict[str, int] = {}
            for col_idx, cell in enumerate(row):
                cell_text = normalize_text(str(cell or ""))
                if not cell_text:
                    continue
                for canonical, aliases in HEADER_ALIASES.items():
                    if cell_text in aliases and canonical not in current_map:
                        current_map[canonical] = col_idx
            score = len(current_map)
            if score > best_score:
                best_score = score
                best_idx = idx
                best_map = current_map
        if best_score <= 0:
            return None, {}
        return best_idx, best_map

    def _parse_rows(
        self,
        project_id: str,
        artifact_id: str,
        filename: str,
        sheet_name: str,
        artifact_type: ArtifactType,
        rows: list[list[Any]],
    ) -> list[EvidenceAtom]:
        if not rows:
            return []
        header_idx, header_map = self._detect_header(rows)
        if header_idx is None:
            return []

        atoms: list[EvidenceAtom] = []
        for row_idx in range(header_idx + 1, len(rows)):
            row = rows[row_idx]
            values = self._extract_row_values(row, header_map)
            if all(not str(v).strip() for v in values.values()):
                continue
            atoms.extend(
                self._row_to_atoms(
                    project_id=project_id,
                    artifact_id=artifact_id,
                    filename=filename,
                    sheet_name=sheet_name,
                    artifact_type=artifact_type,
                    row_number=row_idx + 1,
                    header_map=header_map,
                    values=values,
                )
            )
        return atoms

    def _extract_row_values(self, row: list[Any], header_map: dict[str, int]) -> dict[str, str]:
        extracted: dict[str, str] = {}
        for key, idx in header_map.items():
            value = row[idx] if idx < len(row) else ""
            extracted[key] = str(value).strip() if value is not None else ""
        return extracted

    def _row_to_atoms(
        self,
        project_id: str,
        artifact_id: str,
        filename: str,
        sheet_name: str,
        artifact_type: ArtifactType,
        row_number: int,
        header_map: dict[str, int],
        values: dict[str, str],
    ) -> list[EvidenceAtom]:
        part_number = values.get("part_number", "")
        description = values.get("description", "")
        quantity_raw = values.get("quantity", "")
        unit_price = values.get("unit_price", "")
        lead_time = values.get("lead_time", "")

        entity_keys: list[str] = []
        if description:
            entity_keys.append(normalize_entity_key("device", description))
        if part_number:
            entity_keys.append(normalize_entity_key("part", part_number))

        columns = {key: get_column_letter(index + 1) for key, index in header_map.items()}
        source_ref = SourceRef(
            id=stable_id("src", artifact_id, sheet_name, row_number),
            artifact_id=artifact_id,
            artifact_type=artifact_type,
            filename=filename,
            locator={"sheet": sheet_name, "row": row_number, "columns": columns},
            extraction_method="quote_header_mapping",
            parser_version=self.parser_version,
        )

        atoms: list[EvidenceAtom] = []

        def append_atom(atom_type: AtomType, raw_text: str, value: dict[str, Any], confidence: float) -> None:
            atoms.append(
                EvidenceAtom(
                    id=stable_id("atm", project_id, artifact_id, sheet_name, row_number, atom_type.value, raw_text),
                    project_id=project_id,
                    artifact_id=artifact_id,
                    atom_type=atom_type,
                    raw_text=raw_text,
                    normalized_text=normalize_text(raw_text),
                    value=value,
                    entity_keys=entity_keys,
                    source_refs=[source_ref],
                    authority_class=AuthorityClass.vendor_quote,
                    confidence=confidence,
                    review_status=ReviewStatus.auto_accepted,
                    review_flags=[],
                    parser_version=self.parser_version,
                )
            )

        if part_number or description or quantity_raw or unit_price:
            append_atom(
                AtomType.vendor_line_item,
                f"Line item {part_number} {description}".strip(),
                {
                    "part_number": part_number,
                    "description": description,
                    "quantity": quantity_raw,
                    "unit_price": unit_price,
                    "lead_time": lead_time,
                },
                confidence=0.9,
            )
        if quantity_raw:
            append_atom(
                AtomType.quantity,
                f"Quantity {quantity_raw}",
                parse_quantity(quantity_raw),
                confidence=0.9,
            )
        if lead_time:
            append_atom(
                AtomType.constraint,
                f"Lead time {lead_time}",
                {"lead_time": lead_time},
                confidence=0.85,
            )
        return atoms
